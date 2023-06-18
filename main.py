import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QDialogButtonBox, QTableWidgetItem, QVBoxLayout, QLabel, QPushButton, QWidget, QFormLayout, QLineEdit, QCheckBox, QMessageBox, QFileDialog
from PyQt5 import QtCore
from PyQt5.QtCore import QTimer, Qt, QStandardPaths
from PyQt5.QtGui import QPixmap, QFont
from View.Settings import Ui_Settings
from View.StartPage import Ui_Bauxtifier
from View.StartPageMean import Ui_BauxtifierMean
from View.BauxOutputClass import BauxOutputWindow
import openpyxl
from pycel import ExcelCompiler
import json
import pandas as pd
from decimal import Decimal
import shutil
from PyQt5.QtWidgets import QMainWindow, QWidget, QVBoxLayout, QLabel
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure



class LoginWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Вход")

        layout = QFormLayout()

        self.username_label = QLabel("Логин:")
        self.username_input = QLineEdit()
        layout.addRow(self.username_label, self.username_input)

        self.password_label = QLabel("Пароль:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addRow(self.password_label, self.password_input)

        self.login_button = QPushButton("Войти")
        self.login_button.clicked.connect(self.authenticate)
        layout.addRow(self.login_button)

        container = QWidget()
        container.setLayout(layout)

        self.setCentralWidget(container)

    def authenticate(self):
        if self.password_input.text() == "" and self.username_input.text() == "":
            start_page.show()
            self.close()

from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QLabel, QCheckBox, QSizePolicy, QHBoxLayout
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.start_page1_boxite = StartPage1Boxite(self)
        self.start_page2_boxite = StartPage2Boxite(self)
        self.start_pageMean_boxite = StartPageMeanBoxite(self)
        
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Выбор режима работы")

        btn1 = QPushButton("Расчет материальных \n поток основных переделов ")
        btn1.clicked.connect(self.open_mode1)
        btn2 = QPushButton("Сравнительный анализ бокситов")
        btn2.clicked.connect(self.open_mode2)
        btn3 = QPushButton("Оценка качества боксита\n по усреднённому составу ")
        btn3.clicked.connect(self.open_mode3)
        btn4 = QPushButton("Загрузить данные о бокситах в формате Excel")
        btn4.clicked.connect(self.open_new_window)
        
        btn4.setStyleSheet("font:20px;"
                           "padding : 20px;"
                           "margin-bottom : 10px;"
                           "margin-left : 20px;"
                           "margin-right : 20px;")
        
        btn1.setStyleSheet("font:20px;"
                           "padding : 20px;"
                           "margin-bottom : 10px;"
                           "margin-left : 20px;"
                           "margin-right : 20px;")
        btn2.setStyleSheet("font:20px;"
                           "padding : 20px;"
                           "margin-bottom : 10px;"
                           "margin-left : 20px;"
                           "margin-right : 20px;")
        btn3.setStyleSheet("font:20px;"
                           "padding : 20px;"
                           "margin-left : 20px;"
                           "margin-right : 20px;")
        
        description_label = QLabel("Данная программа автоматически рассчитывает массовые балансы на разных стадиях производства глинозема.")
        description_label.setAlignment(Qt.AlignCenter)
        description_label.setWordWrap(True)

        dark_mode_checkbox = QCheckBox("Темный режим")
        dark_mode_checkbox.toggled.connect(self.toggle_dark_mode)

        dark_mode_layout = QHBoxLayout()
        dark_mode_layout.addStretch(1)
        dark_mode_layout.addWidget(dark_mode_checkbox)

        layout = QVBoxLayout()
        layout.addWidget(description_label)
        layout.addWidget(btn1)
        layout.addWidget(btn2)
        layout.addWidget(btn3)
        layout.addWidget(btn4)
        layout.addLayout(dark_mode_layout)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(20)

        container = QWidget()   
        container.setLayout(layout)

        self.setCentralWidget(container)

        font = QFont()
        font.setPointSize(11)
        font.setBold(True)
        self.setFont(font)

    def open_mode1(self):
        self.start_page1_boxite.show()
        self.hide()

    def open_mode2(self):
        self.start_page2_boxite.show()
        self.hide()

    def open_mode3(self):
        self.start_pageMean_boxite.show()
        self.hide()
        
    def open_new_window(self):
        self.new_window = NewWindow(self, self.start_page1_boxite, self.start_page2_boxite)
        self.new_window.show()
        self.hide()  
        
    def toggle_dark_mode(self, checked):
        if checked:
            app.setStyleSheet(qdarkstyle.load_stylesheet(qdarkstyle.DarkPalette))
        else:
            app.setStyleSheet(qdarkstyle.load_stylesheet(qdarkstyle.LightPalette))

class NewWindow(QMainWindow):

    def __init__(self, main_window, start_page1_boxite, start_page2_boxite):
        super().__init__()

        self.start_page1_boxite = start_page1_boxite
        self.start_page2_boxite = start_page2_boxite
        
        self.main_window = main_window
        self.setWindowTitle("Выбор загрузки")

        btn1 = QPushButton("Расчет материальных \n поток основных переделов")
        btn1.clicked.connect(self.load_file)
        btn2 = QPushButton("Сравнительный анализ бокситов")
        btn2.clicked.connect(self.load_two_files)

        btn1.setStyleSheet("font:20px;"
                           "padding : 20px;"
                           "margin-bottom : 10px;"
                           "margin-left : 20px;"
                           "margin-right : 20px;")
        btn2.setStyleSheet("font:20px;"
                           "padding : 20px;"
                           "margin-bottom : 10px;"
                           "margin-left : 20px;"
                           "margin-right : 20px;")

        layout = QVBoxLayout()
        layout.addWidget(btn1)
        layout.addWidget(btn2)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(20)

        container = QWidget()   
        container.setLayout(layout)

        self.setCentralWidget(container)

        font = QFont()
        font.setPointSize(11)
        font.setBold(True)
        self.setFont(font)
        
    def closeEvent(self, event):
        self.main_window.show()  # Show the MainWindow when NewWindow closes
        event.accept()

    def load_file(self):
        initial_dir = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
        target_file_name, _ = QFileDialog.getOpenFileName(self, "Open File", initial_dir + "/")

        if target_file_name:
            destination = "Excels\Производство глинозема1.xlsx"
            shutil.copy(target_file_name, destination)
            
        self.start_page1_boxite.open_baux_output()

    def load_two_files(self):
        for i in range(2):
            initial_dir = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
            target_file_name, _ = QFileDialog.getOpenFileName(self, "Open File", initial_dir + "/")

            if target_file_name:
                destination = f"Excels\Производство глинозема{i+1}.xlsx"
                shutil.copy(target_file_name, destination)
                
        self.start_page2_boxite.open_baux_output()

class StartPage1Boxite(QMainWindow, Ui_Bauxtifier):
    def __init__(self, previos):
        self.previos = previos
        super().__init__()
        self.current_file_name = None
        self.compiler = None
        self.setupUi(self)
        self.tabWidget.setCurrentIndex(0)
        self.tabWidget.setTabVisible(1, False)
        self.settings_window = SettingsWindow(self)
        self.statusBar().setVisible(False)
        
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), "Состав боксита")
        self.label.setText("Введите состав боксита")
        self.settings_window.tabWidget.setTabText(self.tabWidget.indexOf(self.settings_window.tab), "Исходные данные")
        
        self.settingButton.clicked.connect(self.open_settings)
        self.runButton.clicked.connect(self.open_baux_output)

    def check(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for i in ["1", "2"]:
            for key, value in data.items():
                attr_name = key + i

                attr_value = getattr(self, attr_name, None)

                if attr_value is None:
                    attr_value = getattr(self.settings_window, attr_name, None)

                if attr_value is not None:
                    val = attr_value.text().replace(",", ".")
                    if float(val) < 0:
                        # Неправильный ввод: отрицательные значение в данных
                        QMessageBox.warning(self, "Ошибка ввода", "Неправильный ввод: отрицательные значение в данных")
                        return False
                else:
                    print(f"Атрибут '{attr_name}' не найден ни в self, ни в self.settings_window")

        return True
    
    def closeEvent(self, event):
        self.previos.show()  # Show the MainWindow when NewWindow closes
        event.accept()
        
    def check2(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        bauxite_elements = [
            "al2o3_in_bauxite",
            "fe2o3_in_bauxite",
            "sio2_in_bauxite",
            "tio2_in_bauxite",
            "cao_in_bauxite",
            "co2_in_bauxite",
            "other_elements_in_bauxite",
            "trace_elements_in_bauxite"
        ]

        soda_elements = [
            "soda_moisture",
            "na2co3_in_soda",
            "soda_other_elements"
        ]

        white_sludge_elements = [
            "co2_in_white_sludge",
            "na2o_in_white_sludge",
            "sio2_in_white_sludge",
            "trace_elements_in_white_sludge",
            "al2o3_in_white_sludge",
            "fe2o3_in_white_sludge",
            "cao_in_white_sludge"
        ]

        limestone_elements = [
            "limestone_co2",
            "limestone_sio2",
            "limestone_other_elements",
            "limestone_cao"
        ]

        for i in ["1", "2"]:
            lst = [Decimal(getattr(self, element + i, "0")
                           .text().replace(",", ".")) 
                   for element in bauxite_elements]
            bauxite_sum = sum(lst)
            soda_sum = sum(Decimal(getattr(self, element + i, "0")
                                   .text().replace(",", "."))
                           for element in soda_elements)
            white_sludge_sum = sum(
                Decimal(getattr(self, element + i, "0")
                        .text().replace(",", ".")) 
                for element in white_sludge_elements)
            limestone_sum = sum(
                Decimal(getattr(self, element + i, "0").text()
                        .replace(",", ".")) 
                for element in limestone_elements)
            if bauxite_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", 
                f"Сумма элементов во вкладке Боксит {i} не равна 100%")
                return False
            if soda_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", 
                f"Сумма элементов во вкладке Сода {i} не равна 100%")
                return False
            if white_sludge_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", 
                f"Сумма элементов во вкладке Белый шлам {i} не равна 100%")
                return False
            if limestone_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", 
                f"Сумма элементов во вкладке Известняк {i} не равна 100%")
                return False

        return True

    def open_settings(self):
        self.settings_window.tabWidget.setTabVisible(1, False)
        self.settings_window.show()
        self.setDisabled(True)

    def open_baux_output(self):
        if self.check() and self.check2():
            self.baux_output_window = BauxOutputWindow(self)
            self.baux_output_window.tabWidget_6.setCurrentIndex(0)
            self.baux_output_window.tabWidget.setCurrentIndex(0)
            self.Calculate()
            self.baux_output_window.add_histogram_tab1()
            self.baux_output_window.tabWidget_5.setTabVisible(1, False)
            # self.baux_output_window.tabWidget_5.setTabVisible(2, False)
            self.baux_output_window.show()
            self.hide()

    def table(self, num):
        # Чтение данных из JSON-файла
        with open('./JSON/table.json', 'r', encoding="utf8") as file:
            table_data = json.load(file)

        # Получение данных для таблицы 13
        for i, tab in table_data.items():
        # Считывание таблицы с помощью функции read_table_from_worksheet()
            df = self.read_table_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                tab["start_cell"],
                int(tab["num_rows"]),
                int(tab["num_cols"])
            )
            
            overall = self.read_overall_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                tab["start_cell"],
                int(tab["num_rows"]),
                int(tab["num_cols"])
            )
            # Копирование значений из pandas Series в QTableWidget
            # Здесь предполагается, что у вас уже есть функция copy_series_to_qtablewidget()
            # и экземпляр QTableWidget с именем baux_output_window.tableAluminateSolutionStageOne21
            add = ""
            if num != 1:
                add = f"_{num}"
            for ser, qttab in tab["rows"].items():
                
                self.copy_series_to_qtablewidget(
                    df.loc[ser],
                    getattr(self.baux_output_window, qttab+add)
                )
            
            tab25 = self.read_table_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                "C183",
                7,
                5
            )
            self.table25(
                    tab25,
                    getattr(self.baux_output_window, "tableAlkalineSolution25"+add)
            )
            getattr(self.baux_output_window, "Get"+i+add).setText(f"Всего: {round(overall,2)} кг")
            getattr(self.baux_output_window, "Set"+i+add).setText(f"Всего: {round(overall,2)} кг")
                
    def table25(self, series, table_widget):

        for i in range(series.shape[0]):
            for j in range(series.shape[1]):
                value = series.iloc[i,j]
                item = QTableWidgetItem(str(round(value,2)))
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                table_widget.setItem(i, j, item)
        
    def Calculate(self):
        self.set_data_worksheet()
        self.table(1)
        # self.table(2)
        
    def set_data_worksheet(self):
        # Открываем рабочую книгу
        workbook = openpyxl.load_workbook('Excels/Производство глинозема.xlsx')

        # Выбираем активный лист
        worksheet = workbook["Исходные данные"]
        
        
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for i in ["1"]:
            for key, value in data.items():
                attr_name = key + i
                
                # Попытка найти атрибут в self
                attr_value = getattr(self, attr_name, None)

                # Если атрибут не найден в self, ищем его в self.settings_window
                if attr_value is None:
                    attr_value = getattr(self.settings_window, attr_name, None)

                if attr_value is not None:
                    val = attr_value.text().replace(",", ".")
                    # Запись значения атрибута в ячейку листа
                    if float(val) >= 0:
                        worksheet[value] = float(val)
                    else:
                        print(f"Значение атрибута '{attr_name}' не может быть отрицательным")
                else:
                    print(f"Атрибут '{attr_name}' не найден ни в self, ни в self.settings_window")

            # Сохраняем изменения в рабочей книге и закрываем ее
            workbook.save(f'Excels\\Производство глинозема{i}.xlsx')
            workbook.close()

        
    def copy_series_to_qtablewidget(self, series, table_widget):
        series = series.loc[series != 0].copy()
        if table_widget.rowCount() != 1:
            print("QTableWidget должен содержать только одну строку.")
            return

        if table_widget.columnCount() != 1:
            if table_widget.columnCount() != len(series):
                print("Количество столбцов в QTableWidget и pandas Series должно совпадать.")
                return

        for i, value in enumerate(series):
            item = QTableWidgetItem(str(round(value,2)))
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            table_widget.setItem(0, i, item)
    
    def read_table_from_worksheet(self, file_name, sheet_name="Таблицы", start_cell="A1", num_rows=1, num_cols=1):
        # Создание компилятора Excel
        if file_name != self.current_file_name:
            self.current_file_name = file_name
            self.compiler = ExcelCompiler(filename=file_name)

        # Определение начальной координаты
        start_col_letter = start_cell[0]
        start_row = int(start_cell[1:])

        # Чтение данных из таблицы
        table = []
        for row in range(start_row, start_row + num_rows):
            table_row = []
            for col in range(openpyxl.utils.column_index_from_string(start_col_letter), openpyxl.utils.column_index_from_string(start_col_letter) + num_cols):
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_value = self.compiler.evaluate(f'{sheet_name}!{col_letter}{row}')
                table_row.append(cell_value)
            table.append(table_row)

        # Конвертирование таблицы в pandas DataFrame
        df = pd.DataFrame(table)

        # Установка первой строки как названия колонн и первой колонки как индекс
        df.columns = df.iloc[0].fillna("")
        df.index = df.iloc[:, 0].fillna("1")
        df = df.drop(df.index[0]).drop(df.columns[0], axis=1)
        df = df.dropna(how='all')

        return df
    
    def read_overall_from_worksheet(self, file_name, sheet_name="Таблицы", start_cell="A1", num_rows=1, num_cols=1):
        # Создание компилятора Excel
        if file_name != self.current_file_name:
            self.current_file_name = file_name
            self.compiler = ExcelCompiler(filename=file_name)

        # Определение начальной координаты
        col_letter = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(start_cell[0]) + num_cols - 1)
        row = int(start_cell[1:]) + num_rows - 1
        
        return self.compiler.evaluate(f'{sheet_name}!{col_letter}{row}')

class StartPage2Boxite(QMainWindow, Ui_Bauxtifier):
        
    def check(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for i in ["1", "2"]:
            for key, value in data.items():
                attr_name = key + i

                attr_value = getattr(self, attr_name, None)

                if attr_value is None:
                    attr_value = getattr(self.settings_window, attr_name, None)

                if attr_value is not None:
                    val = attr_value.text().replace(",", ".")
                    if float(val) < 0:
                        # Неправильный ввод: отрицательные значение в данных
                        QMessageBox.warning(self, "Ошибка ввода", "Неправильный ввод: отрицательные значение в данных")
                        return False
                else:
                    print(f"Атрибут '{attr_name}' не найден ни в self, ни в self.settings_window")
        return True
    
    def closeEvent(self, event):
        self.previos.show()  # Show the MainWindow when NewWindow closes
        event.accept()
    
    def check2(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        bauxite_elements = [
            "al2o3_in_bauxite",
            "fe2o3_in_bauxite",
            "sio2_in_bauxite",
            "tio2_in_bauxite",
            "cao_in_bauxite",
            "co2_in_bauxite",
            "other_elements_in_bauxite",
            "trace_elements_in_bauxite"
        ]

        soda_elements = [
            "soda_moisture",
            "na2co3_in_soda",
            "soda_other_elements"
        ]

        white_sludge_elements = [
            "co2_in_white_sludge",
            "na2o_in_white_sludge",
            "sio2_in_white_sludge",
            "trace_elements_in_white_sludge",
            "al2o3_in_white_sludge",
            "fe2o3_in_white_sludge",
            "cao_in_white_sludge"
        ]

        limestone_elements = [
            "limestone_co2",
            "limestone_sio2",
            "limestone_other_elements",
            "limestone_cao"
        ]

        for i in ["1", "2"]:
            bauxite_sum = sum(
                Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in bauxite_elements)
            soda_sum = sum(Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in soda_elements)
            white_sludge_sum = sum(
                Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in white_sludge_elements)
            limestone_sum = sum(
                Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in limestone_elements)

            if bauxite_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Боксит {i} не равна 100%")
                return False

            if soda_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Сода {i} не равна 100%")
                return False

            if white_sludge_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Белый шлам {i} не равна 100%")
                return False

            if limestone_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Известняк {i} не равна 100%")
                return False

        return True

    def __init__(self, previos):
        super().__init__()
        self.current_file_name = None
        self.compiler = None
        self.setupUi(self)
        self.tabWidget.setCurrentIndex(0)
        self.settings_window = SettingsWindow(self)

        self.settingButton.clicked.connect(self.open_settings)
        self.runButton.clicked.connect(self.open_baux_output)

    def check(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for i in ["1", "2"]:
            for key, value in data.items():
                attr_name = key + i

                attr_value = getattr(self, attr_name, None)

                if attr_value is None:
                    attr_value = getattr(self.settings_window, attr_name, None)

                if attr_value is not None:
                    val = attr_value.text().replace(",", ".")
                    if float(val) < 0:
                        # Неправильный ввод: отрицательные значение в данных
                        QMessageBox.warning(self, "Ошибка ввода", "Неправильный ввод: отрицательные значение в данных")
                        return False
                else:
                    print(f"Атрибут '{attr_name}' не найден ни в self, ни в self.settings_window")

        return True
    def open_settings(self):
        self.settings_window.show()
        self.setDisabled(True)

    def open_baux_output(self):
        if self.check() and self.check2():
            self.baux_output_window = BauxOutputWindow(self)
            self.baux_output_window.tabWidget_6.setCurrentIndex(0)
            self.baux_output_window.tabWidget.setCurrentIndex(0)
            self.Calculate()
            self.baux_output_window.add_histogram_tab2()
            self.baux_output_window.show()
            self.hide()

    def table(self, num):
        with open('./JSON/table.json', 'r', encoding="utf8") as file:
            table_data = json.load(file)

        # Получение данных для таблицы 13
        for i, tab in table_data.items():
            # Считывание таблицы с помощью функции read_table_from_worksheet()
            df = self.read_table_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                tab["start_cell"],
                int(tab["num_rows"]),
                int(tab["num_cols"])
            )

            overall = self.read_overall_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                tab["start_cell"],
                int(tab["num_rows"]),
                int(tab["num_cols"])
            )
            # Копирование значений из pandas Series в QTableWidget
            # Здесь предполагается, что у вас уже есть функция copy_series_to_qtablewidget()
            # и экземпляр QTableWidget с именем baux_output_window.tableAluminateSolutionStageOne21
            add = ""
            if num != 1:
                add = f"_{num}"
            for ser, qttab in tab["rows"].items():
                self.copy_series_to_qtablewidget(
                    df.loc[ser],
                    getattr(self.baux_output_window, qttab + add)
                )

            tab25 = self.read_table_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                "C183",
                7,
                5
            )
            self.table25(
                tab25,
                getattr(self.baux_output_window, "tableAlkalineSolution25" + add)
            )

            getattr(self.baux_output_window, "Get" + i + add).setText(f"Всего: {round(overall, 2)} кг")
            # TODO "Set"+i+add переменная для получения всего
            getattr(self.baux_output_window, "Set" + i + add).setText(f"Всего: {round(overall, 2)} кг")

            
    def table25(self, series, table_widget):

        for i in range(series.shape[0]):
            for j in range(series.shape[1]):
                value = series.iloc[i,j]
                item = QTableWidgetItem(str(round(value,2)))
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                table_widget.setItem(i, j, item)
        
    def Calculate(self):
        self.set_data_worksheet()
        self.table(1)
        self.table(2)
        
    def set_data_worksheet(self):
        # Открываем рабочую книгу
        workbook = openpyxl.load_workbook('Excels/Производство глинозема.xlsx')

        # Выбираем активный лист
        worksheet = workbook["Исходные данные"]
        
        
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for i in ["1", "2"]:
            for key, value in data.items():
                attr_name = key + i
                
                # Попытка найти атрибут в self
                attr_value = getattr(self, attr_name, None)

                # Если атрибут не найден в self, ищем его в self.settings_window
                if attr_value is None:
                    attr_value = getattr(self.settings_window, attr_name, None)

                if attr_value is not None:
                    val = attr_value.text().replace(",", ".")
                    # Запись значения атрибута в ячейку листа
                    worksheet[value] = float(val)
                else:
                    print(f"Атрибут '{attr_name}' не найден ни в self, ни в self.settings_window")

            # Сохраняем изменения в рабочей книге и закрываем ее
            workbook.save(f'Excels\\Производство глинозема{i}.xlsx')
            workbook.close()

        
    def copy_series_to_qtablewidget(self, series, table_widget):
        series = series.loc[series != 0].copy()
        if table_widget.rowCount() != 1:
            print("QTableWidget должен содержать только одну строку.")
            return

        if table_widget.columnCount() != 1:
            if table_widget.columnCount() != len(series):
                print("Количество столбцов в QTableWidget и pandas Series должно совпадать.")
                return

        for i, value in enumerate(series):
            item = QTableWidgetItem(str(round(value,2)))
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            table_widget.setItem(0, i, item)
    
    def read_table_from_worksheet(self, file_name, sheet_name="Таблицы", start_cell="A1", num_rows=1, num_cols=1):
        # Создание компилятора Excel
        if file_name != self.current_file_name:
            self.current_file_name = file_name
            self.compiler = ExcelCompiler(filename=file_name)

        # Определение начальной координаты
        start_col_letter = start_cell[0]
        start_row = int(start_cell[1:])

        # Чтение данных из таблицы
        table = []
        for row in range(start_row, start_row + num_rows):
            table_row = []
            for col in range(openpyxl.utils.column_index_from_string(start_col_letter), openpyxl.utils.column_index_from_string(start_col_letter) + num_cols):
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_value = self.compiler.evaluate(f'{sheet_name}!{col_letter}{row}')
                table_row.append(cell_value)
            table.append(table_row)

        # Конвертирование таблицы в pandas DataFrame
        df = pd.DataFrame(table)

        # Установка первой строки как названия колонн и первой колонки как индекс
        df.columns = df.iloc[0].fillna("")
        df.index = df.iloc[:, 0].fillna("1")
        df = df.drop(df.index[0]).drop(df.columns[0], axis=1)
        df = df.dropna(how='all')

        return df
    
    def read_overall_from_worksheet(self, file_name, sheet_name="Таблицы", start_cell="A1", num_rows=1, num_cols=1):
        # Создание компилятора Excel
        if file_name != self.current_file_name:
            self.current_file_name = file_name
            self.compiler = ExcelCompiler(filename=file_name)

        # Определение начальной координаты
        col_letter = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(start_cell[0]) + num_cols - 1)
        row = int(start_cell[1:]) + num_rows - 1
        
        return self.compiler.evaluate(f'{sheet_name}!{col_letter}{row}')

class StartPageMeanBoxite(QMainWindow, Ui_BauxtifierMean):
    def __init__(self, previos):
        self.previos = previos
        super().__init__()
        self.current_file_name = None
        self.compiler = None
        self.setupUi(self)
        self.settings_window = SettingsWindow(self)

        self.settings_window.tabWidget.setTabText(self.settings_window.tabWidget.indexOf(self.settings_window.tab), "Исходные данные")
        
        self.settingButton.clicked.connect(self.open_settings)
        self.runButton.clicked.connect(self.open_baux_output)

    def closeEvent(self, event):
        self.previos.show()  # Show the MainWindow when NewWindow closes
        event.accept()
    
    def check(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for i in ["1"]:
            for key, value in data.items():
                attr_name = key + i

                attr_value = getattr(self, attr_name, None)

                if attr_value is None:
                    attr_value = getattr(self.settings_window, attr_name, None)

                if attr_value is not None:
                    val = attr_value.text().replace(",", ".")
                    if float(val) < 0:
                        # Неправильный ввод: отрицательные значение в данных
                        QMessageBox.warning(self, "Ошибка ввода", "Неправильный ввод: отрицательные значение в данных")
                        return False
                else:
                    print(f"Атрибут '{attr_name}' не найден ни в self, ни в self.settings_window")

        return True
    
    def check2(self):
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        bauxite_elements = [
            "al2o3_in_bauxite",
            "fe2o3_in_bauxite",
            "sio2_in_bauxite",
            "tio2_in_bauxite",
            "cao_in_bauxite",
            "co2_in_bauxite",
            "other_elements_in_bauxite",
            "trace_elements_in_bauxite"
        ]

        soda_elements = [
            "soda_moisture",
            "na2co3_in_soda",
            "soda_other_elements"
        ]

        white_sludge_elements = [
            "co2_in_white_sludge",
            "na2o_in_white_sludge",
            "sio2_in_white_sludge",
            "trace_elements_in_white_sludge",
            "al2o3_in_white_sludge",
            "fe2o3_in_white_sludge",
            "cao_in_white_sludge"
        ]

        limestone_elements = [
            "limestone_co2",
            "limestone_sio2",
            "limestone_other_elements",
            "limestone_cao"
        ]

        for i in ["1"]:
            bauxite_sum = sum(
                Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in bauxite_elements)
            soda_sum = sum(Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in soda_elements)
            white_sludge_sum = sum(
                Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in white_sludge_elements)
            limestone_sum = sum(
                Decimal(getattr(self, element + i, "0").text().replace(",", ".")) for element in limestone_elements)

            if bauxite_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Боксит {i} не равна 100%")
                return False

            if soda_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Сода {i} не равна 100%")
                return False

            if white_sludge_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Белый шлам {i} не равна 100%")
                return False

            if limestone_sum != 100:
                QMessageBox.warning(self, "Ошибка ввода", f"Сумма элементов во вкладке Известняк {i} не равна 100%")
                return False

        return True
    
    def open_settings(self):
        self.setDisabled(True)
        self.settings_window.tabWidget.setTabVisible(1, False)
        self.settings_window.exec_()


    def open_baux_output(self):
        if self.check() and self.check2():
            self.baux_output_window = BauxOutputWindow(self)
            self.baux_output_window.tabWidget_6.setCurrentIndex(0)
            self.baux_output_window.tabWidget.setCurrentIndex(0)
            self.Calculate()
            self.baux_output_window.add_histogram_tab1()
            self.baux_output_window.tabWidget_5.setTabVisible(1, False)
            self.baux_output_window.show()
            self.hide()
        
    def table(self, num):
        # Чтение данных из JSON-файла
        with open('./JSON/table.json', 'r', encoding="utf8") as file:
            table_data = json.load(file)

        # Получение данных для таблицы 13
        for i, tab in table_data.items():
            
        # Считывание таблицы с помощью функции read_table_from_worksheet()
            df = self.read_table_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                tab["start_cell"],
                int(tab["num_rows"]),
                int(tab["num_cols"])
            )
            
            overall = self.read_overall_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                tab["start_cell"],
                int(tab["num_rows"]),
                int(tab["num_cols"])
            )
            # Копирование значений из pandas Series в QTableWidget
            # Здесь предполагается, что у вас уже есть функция copy_series_to_qtablewidget()
            # и экземпляр QTableWidget с именем baux_output_window.tableAluminateSolutionStageOne21
            add = ""
            if num != 1:
                add = f"_{num}"
            for ser, qttab in tab["rows"].items():
                self.copy_series_to_qtablewidget(
                    df.loc[ser],
                    getattr(self.baux_output_window, qttab+add)
                )
            
            tab25 = self.read_table_from_worksheet(
                f"Excels\\Производство глинозема{num}.xlsx",
                "Таблицы",
                "C183",
                7,
                5
            )
            self.table25(
                    tab25,
                    getattr(self.baux_output_window, "tableAlkalineSolution25"+add)
            )
            getattr(self.baux_output_window, "Get"+i+add).setText(f"Всего: {round(overall,2)} кг")
            getattr(self.baux_output_window, "Set"+i+add).setText(f"Всего: {round(overall,2)} кг")
                
    def table25(self, series, table_widget):

        for i in range(series.shape[0]):
            for j in range(series.shape[1]):
                value = series.iloc[i,j]
                item = QTableWidgetItem(str(round(value,2)))
                item.setTextAlignment(QtCore.Qt.AlignCenter)
                table_widget.setItem(i, j, item)
        
    def Calculate(self):
        self.set_data_worksheet()
        self.table(1)
        
    def set_data_worksheet(self):
        # Открываем рабочую книгу
        workbook = openpyxl.load_workbook('Excels/Производство глинозема.xlsx')

        # Выбираем активный лист
        worksheet = workbook["Исходные данные"]
        
        
        with open("./JSON/data.json", "r") as file:
            data = json.load(file)

        for key, value in data.items():
            if "bauxite" in key:
                attr_name1 = key + "1"
                attr_name2 = key + "2"
                attr_value1 = getattr(self, attr_name1, None)
                attr_value2 = getattr(self, attr_name2, None)
                if attr_value1 is None:
                    attr_valueMean = getattr(self.settings_window, attr_name1, None)
                    attr_valueMean = float(attr_valueMean.text().replace(",", "."))
                else:
                    attr_value1 = float(attr_value1.text().replace(",", "."))
                    attr_value2 = float(attr_value2.text().replace(",", "."))
                    
                    self.per = float(self.percentage.text())/100
                    
                    attr_valueMean = (attr_value1 * (self.per) + attr_value2 * (1 - self.per))
            else:
                attr_name1 = key + "1"
                attr_value1 = getattr(self, attr_name1, None)
                if attr_value1 is None:
                    attr_valueMean = getattr(self.settings_window, attr_name1, None)
                else:
                    attr_valueMean = attr_value1
                attr_valueMean = float(attr_valueMean.text().replace(",", "."))
            
            # Попытка найти атрибут в self
            
            # Если атрибут не найден в self, ищем его в self.settings_window
            
            
            
            if attr_valueMean is not None:
                worksheet[value] = attr_valueMean
            else:
                print(f"Атрибут '{attr_name1}' не найден ни в self, ни в self.settings_window")
                print(f"Атрибут '{attr_name2}' не найден ни в self, ни в self.settings_window")

        # Сохраняем изменения в рабочей книге и закрываем ее
        
        workbook.save(f'Excels\\Производство глинозема1.xlsx')
        workbook.close()

    def copy_series_to_qtablewidget(self, series, table_widget):
        series = series.loc[series != 0].copy()
        if table_widget.rowCount() != 1:
            print("QTableWidget должен содержать только одну строку.")
            return

        if table_widget.columnCount() != 1:
            if table_widget.columnCount() != len(series):
                print("Количество столбцов в QTableWidget и pandas Series должно совпадать.")
                return

        for i, value in enumerate(series):
            item = QTableWidgetItem(str(round(value,2)))
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            table_widget.setItem(0, i, item)
    
    def read_table_from_worksheet(self, file_name, sheet_name="Таблицы", start_cell="A1", num_rows=1, num_cols=1):
        # Создание компилятора Excel
        if file_name != self.current_file_name:
            self.current_file_name = file_name
            self.compiler = ExcelCompiler(filename=file_name)

        # Определение начальной координаты
        start_col_letter = start_cell[0]
        start_row = int(start_cell[1:])

        # Чтение данных из таблицы
        table = []
        for row in range(start_row, start_row + num_rows):
            table_row = []
            for col in range(openpyxl.utils.column_index_from_string(start_col_letter), openpyxl.utils.column_index_from_string(start_col_letter) + num_cols):
                col_letter = openpyxl.utils.get_column_letter(col)
                cell_value = self.compiler.evaluate(f'{sheet_name}!{col_letter}{row}')
                table_row.append(cell_value)
            table.append(table_row)

        # Конвертирование таблицы в pandas DataFrame
        df = pd.DataFrame(table)

        # Установка первой строки как названия колонн и первой колонки как индекс
        df.columns = df.iloc[0].fillna("")
        df.index = df.iloc[:, 0].fillna("1")
        df = df.drop(df.index[0]).drop(df.columns[0], axis=1)
        df = df.dropna(how='all')

        return df
    
    def read_overall_from_worksheet(self, file_name, sheet_name="Таблицы", start_cell="A1", num_rows=1, num_cols=1):
        # Создание компилятора Excel
        if file_name != self.current_file_name:
            self.current_file_name = file_name
            self.compiler = ExcelCompiler(filename=file_name)

        # Определение начальной координаты
        col_letter = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(start_cell[0]) + num_cols - 1)
        row = int(start_cell[1:]) + num_rows - 1
        
        return self.compiler.evaluate(f'{sheet_name}!{col_letter}{row}')


class SettingsWindow(QDialog, Ui_Settings):
    def __init__(self, start_page):
        super().__init__()
        self.setupUi(self)
        self.tabWidget.setCurrentIndex(0)
        self.start_page = start_page
        self.adjustSize()
        
        button_save = self.buttonBox.button(QDialogButtonBox.Save)
        button_save.setText("Сохранить")
        
        button_discard = self.buttonBox.button(QDialogButtonBox.Discard)
        button_discard.setText("Отменить")
        
        button_restore_defaults = self.buttonBox.button(QDialogButtonBox.RestoreDefaults)
        button_restore_defaults.setText("Восстановить по умолчанию")

        # Подключаем слоты к сигналу clicked
        button_save.clicked.connect(self.save_settings)
        button_discard.clicked.connect(self.discard_changes)
        button_restore_defaults.clicked.connect(self.restore_defaults)

    def closeEvent(self, event):
        self.start_page.setEnabled(True)
        event.accept()
        
    # Функции для кнопок
    def save_settings(self):
        self.start_page.setEnabled(True)

    def discard_changes(self):
        self.start_page.setEnabled(True)
        self.close()
        pass

    def restore_defaults(self):
        # код для восстановления настроек по умолчанию
        pass

class LoadingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("Загрузка")
        self.setWindowModality(Qt.ApplicationModal)
        self.setFixedSize(200, 100)

        layout = QVBoxLayout()

        self.label = QLabel("Проводятся расчеты", self)
        layout.addWidget(self.label, alignment=Qt.AlignHCenter)

        self.setLayout(layout)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_dots)
        self.timer.start(300)

        self.dot_count = 0

    def update_dots(self):
        self.dot_count = (self.dot_count + 1) % 4
        self.label.setText(f"Проводятся расчеты{'.' * self.dot_count}")

import qdarkstyle

if __name__ == '__main__':
    app = QApplication(sys.argv)
    font = QFont()
    font.setPointSize(11)
    app.setFont(font)
    app.setStyleSheet(qdarkstyle.load_stylesheet(qdarkstyle.LightPalette))
    login_window = LoginWindow()
    start_page = MainWindow()
    login_window.show()
    sys.exit(app.exec_())
